"""
Casting Log Parser — Ingestion & Cleaning of 'Actual Casting Log 2025.xlsx'.

Features:
1. High-speed streaming parser (openpyxl read_only with automatic empty-row cutoff).
2. Normalizes dates (e.g. '2025-02-12' from datetime objects or strings).
3. Extracts actual measured weight, allowable job card weight, calculated weight, and weight variance.
4. Extracts tooling, pattern, and molding process metadata:
   - Mold process (Alpha Set, etc.)
   - Core process (Alpha Set, Dolomite, etc.)
   - Riser % & Technology
   - Simulation paths & shaft fitting
   - Size with Contraction Allowance (CA) and core box sizes.
5. Normalizes job numbers, piece numbers, and dimensional parameters.

Output: data/processed/cleaned_casting_log_2025.json
"""
import datetime
import json
import logging
import re
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from backend.services.normalizer import (
    normalize_job_number, normalize_piece_number, extract_base_job
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

RAW_DIR = PROJECT_ROOT / "data" / "raw"
OUTPUT_DIR = PROJECT_ROOT / "data" / "processed"
OUTPUT_FILE = OUTPUT_DIR / "cleaned_casting_log_2025.json"

CASTING_LOG_CANDIDATES = [
    RAW_DIR / "Actual Casting Log 2025.xlsx",
    Path(r"d:\ML\Qadri ML project\Data For Project (Mill Roller Shell Data based)") / "Actual Casting Log 2025.xlsx",
]


def safe_float(val) -> float | None:
    """Convert cell value to float or return None."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if "/" in s and all(p.strip().isdigit() for p in s.split("/")):
        parts = s.split("/")
        if len(parts) == 2:
            try:
                return float(parts[0]) / float(parts[1])
            except (ValueError, ZeroDivisionError):
                pass
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def safe_int(val) -> int | None:
    """Convert cell value to int or return None."""
    f = safe_float(val)
    return int(round(f)) if f is not None else None


def safe_str(val) -> str | None:
    """Clean string cell value."""
    if val is None or val == "" or str(val).strip() == "None":
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        try:
            float(s)
            s = s[:-2]
        except ValueError:
            pass
    return s if s else None


def format_date_str(val) -> str | None:
    """Normalize date value to YYYY-MM-DD string."""
    if val is None or val == "" or str(val).strip() == "None":
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Check for YYYY-MM-DD pattern
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # Check for DD-MM-YYYY pattern
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return s if len(s) >= 4 else None


def normalize_month(val: str | None) -> str | None:
    """Normalize month name to standard 3-letter abbreviation."""
    if not val or str(val).strip() == "" or str(val).strip().lower() == "none":
        return None
    v = str(val).strip().lower()
    mapping = {
        "jan": "Jan", "january": "Jan",
        "feb": "Feb", "february": "Feb",
        "mar": "Mar", "march": "Mar",
        "apr": "Apr", "april": "Apr",
        "may": "May",
        "jun": "Jun", "june": "Jun",
        "jul": "Jul", "july": "Jul",
        "aug": "Aug", "august": "Aug",
        "sep": "Sep", "september": "Sep", "sept": "Sep",
        "oct": "Oct", "october": "Oct",
        "nov": "Nov", "november": "Nov",
        "dec": "Dec", "december": "Dec",
    }
    return mapping.get(v, str(val).strip().capitalize())


def normalize_process_name(val: str | None) -> str | None:
    """Standardize molding and core sand process names (e.g. 'Alpha set' -> 'Alpha Set', 'Ms Pipe' -> 'MS Pipe')."""
    if not val or str(val).strip() == "" or str(val).strip().lower() == "none":
        return None
    s = str(val).strip()
    if s.lower() == "alpha set":
        return "Alpha Set"
    if s.lower() == "co2 + chills":
        return "Co2 + Chills"
    if "ms pipe" in s.lower():
        s = re.sub(r"(?i)ms pipe", "MS Pipe", s)
    return s


def calculate_wall_thickness(od: float | None, id_dim: float | None) -> float | None:
    """Compute radial wall thickness = (OD - ID) / 2."""
    if od is not None and id_dim is not None and od > id_dim and id_dim >= 0:
        return round((od - id_dim) / 2.0, 2)
    return None


def parse_casting_log(filepath: Path | str, year: int = 2025) -> list[dict]:
    """
    Parse Actual Casting Log Excel file with streaming iterator.
    Handles .xlsx (openpyxl) and legacy .xls (xlrd) formats, and dynamic multi-sheet configurations.
    """
    path = Path(filepath)
    if not path.exists():
        log.error(f"Casting log not found: {path}")
        return []

    log.info(f"Opening Casting Log: {path.name} (Year: {year})")

    rows = []
    if path.suffix.lower() == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            target_sheet = None
            year_str = str(year)
            for sn in wb.sheet_names():
                if year_str in sn or "casting" in sn.lower() or "log" in sn.lower():
                    target_sheet = sn
                    break
            if not target_sheet:
                target_sheet = wb.sheet_names()[0]
            sheet = wb.sheet_by_name(target_sheet)
            log.info(f"Parsing sheet: '{target_sheet}' via xlrd...")
            for r in range(sheet.nrows):
                rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        except Exception as e:
            log.error(f"Failed to open .xls workbook {path}: {e}")
            return []
    else:
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            target_sheet = None
            year_str = str(year)
            for sn in wb.sheetnames:
                if year_str in sn or "casting" in sn.lower() or "log" in sn.lower():
                    target_sheet = sn
                    break
            if not target_sheet:
                target_sheet = wb.sheetnames[0]
            sheet = wb[target_sheet]
            log.info(f"Parsing sheet: '{target_sheet}'...")
            rows = list(sheet.iter_rows(values_only=True))
        except Exception as e:
            log.error(f"Failed to open workbook {path}: {e}")
            return []

    header_keywords = [
        "job no", "job #", "idm", "sr", "drawing", "finish", "casted",
        "piece", "name", "month", "lot", "od", "id", "length", "size"
    ]

    header_rows_idx = []
    for r_idx in range(min(6, len(rows))):
        row = rows[r_idx]
        row_vals = [str(v or "").strip() for v in row]
        row_str = " ".join(row_vals).lower()
        has_real_job = any(re.match(r"^[A-Za-z0-9]{2,4}-[A-Za-z0-9]{3,5}-\d{4}", v) for v in row_vals if v)
        if has_real_job:
            break
        if any(k in row_str for k in header_keywords):
            header_rows_idx.append(r_idx)
        elif header_rows_idx:
            break

    col_map = {
        "month": 0, "lot_number": 1, "serial": 2, "idm": 3, "job_number": 4,
        "name": 5, "drawing": 6, "shell_type": 7, "piece": 8,
        "finish_od": 9, "finish_id": 10, "finish_length": 11,
        "cast_od": 12, "cast_id": 13, "cast_length": 14,
        "pattern_ca": 15, "core_box": 20, "riser_pct": 24, "simulation_path": 27,
        "calc_weight": 28, "job_card_weight": 34, "weight_diff": 40, "actual_weight": 41,
        "cast_date": 42, "mat_standard": 43, "technology": 44, "shaft_fitting": 45,
        "core_process": 46, "mold_process": 47,
    }

    for hr in header_rows_idx:
        row = rows[hr]
        for c, val in enumerate(row):
            v = str(val or "").strip().replace("\n", " ").lower()
            if "month" in v: col_map["month"] = c
            elif "lot" in v and "wt" not in v: col_map["lot_number"] = c
            elif "sr" in v and "job" not in v: col_map["serial"] = c
            elif "idm" in v: col_map["idm"] = c
            elif "job no" in v or "job #" in v or (v.startswith("job") and "wt" not in v and "card" not in v): col_map["job_number"] = c
            elif (v == "name" or "shell name" in v or "item name" in v) and "drawing" not in v: col_map["name"] = c
            elif "drawing" in v: col_map["drawing"] = c
            elif "type" in v and "shell" in v: col_map["shell_type"] = c
            elif "piece" in v: col_map["piece"] = c
            elif ("actual wt" in v or "act wt" in v or "act. wt" in v or v == "actual") and "cage" not in v: col_map["actual_weight"] = c
            elif "job card wt" in v or "card wt" in v or (v.startswith("wt") and "cage" not in v): col_map["job_card_weight"] = c
            elif "calc" in v and "wt" in v: col_map["calc_weight"] = c
            elif "pattern" in v: col_map["pattern_ca"] = c
            elif "core box" in v: col_map["core_box"] = c
            elif "riser" in v: col_map["riser_pct"] = c

    # Find Finish & Casted dimensional column bounds
    finish_start, cast_start = None, None
    for hr in header_rows_idx:
        row = rows[hr]
        for c, val in enumerate(row):
            v = str(val or "").strip().replace("\n", " ").lower()
            if "finish" in v and finish_start is None: finish_start = c
            elif "cast" in v and cast_start is None: cast_start = c

    for hr in header_rows_idx:
        row = rows[hr]
        r_vals = [str(v or "").strip().replace("\n", " ").lower() for v in row]
        if finish_start is not None:
            end_f = cast_start if cast_start is not None else finish_start + 4
            for c in range(finish_start, min(end_f, len(r_vals))):
                v2 = r_vals[c]
                if v2 == "od": col_map["finish_od"] = c
                elif v2 == "id": col_map["finish_id"] = c
                elif "len" in v2 or "length" in v2: col_map["finish_length"] = c

        if cast_start is not None:
            end_c = cast_start + 4
            for c in range(cast_start, min(end_c, len(r_vals))):
                v2 = r_vals[c]
                if v2 == "od": col_map["cast_od"] = c
                elif v2 == "id": col_map["cast_id"] = c
                elif "len" in v2 or "length" in v2: col_map["cast_length"] = c

    data_start = max(header_rows_idx) + 1 if header_rows_idx else 2
    records = []
    consecutive_empty = 0
    current_lot = None
    current_month = None

    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        if not any(row):
            consecutive_empty += 1
            if consecutive_empty > 20:
                break
            continue

        consecutive_empty = 0

        job_col = col_map.get("job_number", 4)
        job_raw = row[job_col] if len(row) > job_col else None
        job_number = normalize_job_number(job_raw)
        if not job_number or job_number.upper() in ["JOB NO", "JOB #", "JOB", "NONE", "SR", "IDM"]:
            continue

        name_col = col_map.get("name", 5)
        shell_name = safe_str(row[name_col]) if len(row) > name_col else None
        if not shell_name or shell_name.lower() in ["name", "item name", "shell name", "temp"]:
            continue

        # Handle inverted IDM and Job Number columns in irregular sheets
        if "IDM" in job_number.upper() and re.match(r"^[A-Za-z0-9]{2,4}-[A-Za-z0-9]{3,5}-\d{4}", shell_name):
            job_number = shell_name
            shell_name = "Mill Roller Shell"

        # Forward fill lot and month
        lot_col = col_map.get("lot_number", 1)
        raw_lot = safe_int(row[lot_col]) if len(row) > lot_col else None
        if raw_lot is not None and raw_lot > 0:
            current_lot = raw_lot
        lot_number = current_lot

        month_col = col_map.get("month", 0)
        raw_m = normalize_month(safe_str(row[month_col])) if len(row) > month_col else None
        if raw_m:
            current_month = raw_m
        month = current_month

        ser_col = col_map.get("serial", 2)
        serial_number = safe_int(row[ser_col]) if len(row) > ser_col else None
        idm_col = col_map.get("idm", 3)
        idm_number = safe_str(row[idm_col]) if len(row) > idm_col else None
        draw_col = col_map.get("drawing", 6)
        drawing_number = safe_str(row[draw_col]) if len(row) > draw_col else None
        type_col = col_map.get("shell_type", 7)
        shell_type = safe_str(row[type_col]) if len(row) > type_col else None
        pc_col = col_map.get("piece", 8)
        piece_number = normalize_piece_number(safe_str(row[pc_col])) if len(row) > pc_col else None

        # Dimensions
        fod_c = col_map.get("finish_od", 9)
        fid_c = col_map.get("finish_id", 10)
        flen_c = col_map.get("finish_length", 11)
        finish_od = safe_float(row[fod_c]) if len(row) > fod_c else None
        finish_id = safe_float(row[fid_c]) if len(row) > fid_c else None
        finish_len = safe_float(row[flen_c]) if len(row) > flen_c else None

        cod_c = col_map.get("cast_od", 12)
        cid_c = col_map.get("cast_id", 13)
        clen_c = col_map.get("cast_length", 14)
        cast_od = safe_float(row[cod_c]) if len(row) > cod_c else None
        cast_id = safe_float(row[cid_c]) if len(row) > cid_c else None
        cast_len = safe_float(row[clen_c]) if len(row) > clen_c else None

        # Filter out rows with non-shell / template dimensions
        if finish_od is not None and finish_od <= 50:
            finish_od = None
        if cast_od is not None and cast_od <= 50:
            cast_od = None

        wall_thickness = calculate_wall_thickness(finish_od, finish_id)
        cast_wall_thickness = calculate_wall_thickness(cast_od, cast_id)

        # Tooling & Pattern
        pat_c = col_map.get("pattern_ca", 15)
        cb_c = col_map.get("core_box", 20)
        riser_c = col_map.get("riser_pct", 24)
        sim_c = col_map.get("simulation_path", 27)
        pattern_ca = safe_float(row[pat_c]) if len(row) > pat_c else None
        core_box = safe_float(row[cb_c]) if len(row) > cb_c else None
        riser_pct = safe_float(row[riser_c]) if len(row) > riser_c else None
        simulation_path = safe_str(row[sim_c]) if len(row) > sim_c else None

        # Weights (kg)
        calc_c = col_map.get("calc_weight", 28)
        jc_c = col_map.get("job_card_weight", 34)
        wd_c = col_map.get("weight_diff", 40)
        act_c = col_map.get("actual_weight", 41)
        calculated_weight = safe_float(row[calc_c]) if len(row) > calc_c else None
        job_card_weight = safe_float(row[jc_c]) if len(row) > jc_c else None
        weight_diff = safe_float(row[wd_c]) if len(row) > wd_c else None
        actual_weight = safe_float(row[act_c]) if len(row) > act_c else None

        # If weight_diff is not explicitly recorded but actual & job card are present, calculate it
        if weight_diff is None and actual_weight is not None and job_card_weight is not None:
            weight_diff = round(actual_weight - job_card_weight, 2)

        # Casting Date & Foundry Tech
        dt_c = col_map.get("cast_date", 42)
        mat_c = col_map.get("mat_standard", 43)
        tech_c = col_map.get("technology", 44)
        shaft_c = col_map.get("shaft_fitting", 45)
        core_c = col_map.get("core_process", 46)
        mold_c = col_map.get("mold_process", 47)
        cast_date = format_date_str(row[dt_c]) if len(row) > dt_c else None
        mat_standard = safe_str(row[mat_c]) if len(row) > mat_c else None
        technology = safe_str(row[tech_c]) if len(row) > tech_c else None
        shaft_fitting = safe_str(row[shaft_c]) if len(row) > shaft_c else None
        core_process = normalize_process_name(safe_str(row[core_c])) if len(row) > core_c else None
        mold_process = normalize_process_name(safe_str(row[mold_c])) if len(row) > mold_c else None

        record = {
            "month": month,
            "lot_number": lot_number,
            "serial_number": serial_number,
            "idm_number": idm_number,
            "job_number": job_number,
            "base_job": extract_base_job(job_number),
            "shell_name": shell_name,
            "drawing_number": drawing_number,
            "shell_type": shell_type,
            "piece_number": piece_number,
            "od": finish_od,
            "id_dim": finish_id,
            "length": finish_len,
            "wall_thickness": wall_thickness,
            "cast_od": cast_od,
            "cast_id": cast_id,
            "cast_length": cast_len,
            "cast_wall_thickness": cast_wall_thickness,
            "pattern_size_ca": pattern_ca,
            "core_box": core_box,
            "riser_pct": riser_pct,
            "simulation_path": simulation_path,
            "calculated_weight": calculated_weight,
            "job_card_weight": job_card_weight,
            "actual_weight": actual_weight,
            "weight_diff": weight_diff,
            "weight": job_card_weight or calculated_weight or actual_weight,
            "cast_date": cast_date,
            "material_standard": mat_standard,
            "technology": technology,
            "shaft_fitting": shaft_fitting,
            "core_process": core_process,
            "mold_process": mold_process,
            "data_year": year,
            "file_path": str(path),
            "sheet_name": target_sheet,
        }

        # Ignore empty summary rows with zero physical dimensions and zero weight
        if (
            finish_od is None
            and cast_od is None
            and actual_weight is None
            and job_card_weight is None
            and calculated_weight is None
        ):
            continue

        records.append(record)

    log.info(f"Successfully extracted {len(records)} casting log records from {path.name}")
    return records


def find_casting_log_file(custom_path: str | Path | None = None) -> Path | None:
    """Find the casting log file from candidate locations."""
    if custom_path:
        p = Path(custom_path)
        if p.exists():
            return p

    for candidate in CASTING_LOG_CANDIDATES:
        if candidate.exists():
            return candidate

    # Search in raw data dir
    if RAW_DIR.exists():
        for f in RAW_DIR.glob("*.xlsx"):
            if "casting" in f.name.lower():
                return f

    return None


def main():
    """CLI execution for Casting Log extraction."""
    log.info("=" * 60)
    log.info("ACTUAL CASTING LOG PARSER — Starting")
    log.info("=" * 60)

    casting_log_path = find_casting_log_file()

    if casting_log_path is None:
        log.warning(
            "Actual Casting Log 2025.xlsx NOT FOUND in any candidate location.\n"
            "  Checked:\n" +
            "\n".join(f"    - {p}" for p in CASTING_LOG_CANDIDATES)
        )
        records = []
    else:
        log.info(f"Found Casting Log file at: {casting_log_path}")
        records = parse_casting_log(casting_log_path, year=2025)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False, default=str)

    log.info(f"Saved {len(records)} casting records to {OUTPUT_FILE}")
    log.info("ACTUAL CASTING LOG PARSER — Complete")


if __name__ == "__main__":
    main()
