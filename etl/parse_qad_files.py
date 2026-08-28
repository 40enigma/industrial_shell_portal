"""
QAD / QDAR File Parser — Extracts defect logs and quality reports.

Extracts:
- QDAR #, Date, Customer Name, Job #, Part Name, Piece No, Drawing #
- Defect Description & Investigation Remarks (Row 18 / Row 17)
- Defect Judgment (Rework able / Reject / Concession)
- Detected At Stage, Detected By, and Responsible Department
- Auto-fallback between openpyxl and xlrd formats.

Output: data/processed/qdar_mapping.json
"""
import json
import logging
import re
import sys
from pathlib import Path

import openpyxl
import xlrd

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from backend.services.normalizer import normalize_job_number, normalize_piece_number

_workspace_root = PROJECT_ROOT.parent
_candidate_parent = _workspace_root / "Data For Project (Mill Roller Shell Data based)"

def _resolve_qdar_dir(year: int = 2025) -> Path:
    if _candidate_parent.exists():
        for pat in [f"*QDAR*{year}*", f"*QDR*{year}*", f"{year}/*QDAR*", f"{year}/*QDR*"]:
            matches = list(_candidate_parent.glob(pat))
            if matches:
                return matches[0]
        if (_candidate_parent / str(year)).exists():
            return _candidate_parent / str(year)
    return _candidate_parent / f"QDARS {year}"

RAW_QDAR_DIR = _resolve_qdar_dir(2025)
OUTPUT_DIR = PROJECT_ROOT / "data" / "processed"
OUTPUT_FILE = OUTPUT_DIR / "qdar_mapping.json"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None"):
        return None
    return s


# ---------------------------------------------------------------------------
# Cell Coordinates (1-indexed for openpyxl)
# ---------------------------------------------------------------------------

EXTERNAL_CELLS = {
    "qdar_number":        (3, 9),   # I3
    "doc_date":           (4, 9),   # I4
    "detected_by":        (8, 8),   # H8
    "detected_at":        (10, 4),  # D10
    "job_number":         (12, 8),  # H12
    "customer_name":      (12, 4),  # D12
    "part_name":          (13, 4),  # D13
    "drawing_number":     (13, 8),  # H13
    "piece_number":       (14, 4),  # D14
    "responsibility":     (14, 8),  # H14
    "defect_description": (18, 2),  # B18
    "defect_judgment":    (30, 10), # J30
}

INTERNAL_CELLS = {
    "qdar_number":        (3, 9),   # I3
    "doc_date":           (4, 9),   # I4
    "detected_by":        (8, 8),   # H8
    "detected_at":        (9, 4),   # D9
    "job_number":         (11, 8),  # H11
    "customer_name":      (11, 4),  # D11
    "part_name":          (12, 4),  # D12
    "drawing_number":     (12, 8),  # H12
    "piece_number":       (13, 4),  # D13
    "responsibility":     (13, 8),  # H13
    "defect_description": (17, 2),  # B17
    "defect_judgment":    (30, 4),  # D30
}


def _extract_fuzzy_fields_openpyxl(sh, record: dict):
    """Scan top 35 rows for Job Number, Defect Judgment, and Description if coordinates failed."""
    max_r = min(sh.max_row or 35, 40)
    max_c = min(sh.max_column or 15, 15)
    judgment_keywords = ["reject", "rework", "concession", "accept", "use as is"]

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            val = str(sh.cell(row=r, column=c).value or "").strip()
            if not val:
                continue
            # Fuzzy job number extraction
            if not record.get("job_number"):
                m = re.search(r"([A-Z]{1,2}\d{2}-[A-Z0-9]{2,6}-\d{3,5})", val.upper())
                if m:
                    record["job_number"] = m.group(1)
            # Fuzzy judgment extraction
            if not record.get("defect_judgment"):
                val_lower = val.lower()
                for kw in judgment_keywords:
                    if kw in val_lower:
                        record["defect_judgment"] = val
                        break
            # Fuzzy description extraction (look for labels)
            if not record.get("defect_description"):
                val_lower = val.lower()
                if "defect" in val_lower and ("desc" in val_lower or "detail" in val_lower or "nature" in val_lower):
                    # Read the cell to the right or below as the actual description
                    right_val = safe_str(sh.cell(row=r, column=c + 1).value) if c < max_c else None
                    below_val = safe_str(sh.cell(row=r + 1, column=c).value) if r < max_r else None
                    record["defect_description"] = right_val or below_val


def _extract_fuzzy_fields_xlrd(sh, record: dict):
    """Scan top 35 rows for Job Number, Defect Judgment, and Description if coordinates failed."""
    judgment_keywords = ["reject", "rework", "concession", "accept", "use as is"]

    for r in range(min(sh.nrows, 40)):
        for c in range(min(sh.ncols, 15)):
            val = str(sh.cell_value(r, c)).strip()
            if not val:
                continue
            # Fuzzy job number extraction
            if not record.get("job_number"):
                m = re.search(r"([A-Z]{1,2}\d{2}-[A-Z0-9]{2,6}-\d{3,5})", val.upper())
                if m:
                    record["job_number"] = m.group(1)
            # Fuzzy judgment extraction
            if not record.get("defect_judgment"):
                val_lower = val.lower()
                for kw in judgment_keywords:
                    if kw in val_lower:
                        record["defect_judgment"] = val
                        break
            # Fuzzy description extraction
            if not record.get("defect_description"):
                val_lower = val.lower()
                if "defect" in val_lower and ("desc" in val_lower or "detail" in val_lower or "nature" in val_lower):
                    right_val = safe_str(sh.cell_value(r, c + 1)) if c + 1 < sh.ncols else None
                    below_val = safe_str(sh.cell_value(r + 1, c)) if r + 1 < sh.nrows else None
                    record["defect_description"] = right_val or below_val


def parse_xlsx_qdar(filepath: Path, cell_map: dict, doc_subtype: str, year: int = 2025) -> dict | None:
    """Parse .xlsx QDAR workbook using openpyxl."""
    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    except Exception:
        return None

    try:
        sh = wb.active if wb.active is not None else wb[wb.sheetnames[0]]

        record = {
            "file_path": str(filepath),
            "file_name": filepath.name,
            "doc_type": doc_subtype,
            "data_year": year,
        }

        for field, (r, c) in cell_map.items():
            try:
                record[field] = safe_str(sh.cell(row=r, column=c).value)
            except Exception:
                record[field] = None

        _extract_fuzzy_fields_openpyxl(sh, record)

        # Multi-row defect description: read up to 5 rows starting from the coordinate
        desc_coord = cell_map.get("defect_description")
        if desc_coord and record.get("defect_description"):
            desc_r, desc_c = desc_coord
            parts = [record["defect_description"]]
            for extra_r in range(desc_r + 1, min(desc_r + 5, (sh.max_row or desc_r) + 1)):
                extra_val = safe_str(sh.cell(row=extra_r, column=desc_c).value)
                if extra_val and len(extra_val) > 2:
                    parts.append(extra_val)
                else:
                    break
            if len(parts) > 1:
                record["defect_description"] = " | ".join(parts)

        wb.close()

        record["job_number"] = normalize_job_number(record.get("job_number"))
        record["piece_number"] = normalize_piece_number(record.get("piece_number"))
        return record
    except Exception:
        try: wb.close()
        except Exception: pass
        return None


def parse_xls_qdar(filepath: Path, cell_map: dict, doc_subtype: str, year: int = 2025) -> dict | None:
    """Parse legacy .xls QDAR workbook using xlrd."""
    try:
        wb = xlrd.open_workbook(str(filepath))
    except Exception:
        return None

    try:
        sh = wb.sheet_by_index(0)
        record = {
            "file_path": str(filepath),
            "file_name": filepath.name,
            "doc_type": doc_subtype,
            "data_year": year,
        }

        for field, (r_1, c_1) in cell_map.items():
            r_0, c_0 = r_1 - 1, c_1 - 1
            if r_0 < sh.nrows and c_0 < sh.ncols:
                record[field] = safe_str(sh.cell_value(r_0, c_0))
            else:
                record[field] = None

        _extract_fuzzy_fields_xlrd(sh, record)

        record["job_number"] = normalize_job_number(record.get("job_number"))
        record["piece_number"] = normalize_piece_number(record.get("piece_number"))
        return record
    except Exception:
        return None


def parse_qdar_directory(dirpath: Path, doc_subtype: str, year: int = 2025) -> list[dict]:
    """Parse workbooks in a directory with xlrd/openpyxl fallback."""
    records = []
    if not dirpath.exists():
        return records

    cell_map = INTERNAL_CELLS if doc_subtype == "QDR_INTERNAL" else EXTERNAL_CELLS
    files = sorted(dirpath.iterdir())
    xlsx_files = [f for f in files if f.suffix.lower() in (".xlsx", ".xls") and not f.name.startswith("~$")]

    log.info(f"  Scanning {dirpath.name}: {len(xlsx_files)} workbooks")

    for fpath in xlsx_files:
        rec = None
        if fpath.suffix.lower() == ".xls":
            rec = parse_xls_qdar(fpath, cell_map, doc_subtype, year=year)
            if rec is None:
                rec = parse_xlsx_qdar(fpath, cell_map, doc_subtype, year=year)
        else:
            rec = parse_xlsx_qdar(fpath, cell_map, doc_subtype, year=year)
            if rec is None:
                rec = parse_xls_qdar(fpath, cell_map, doc_subtype, year=year)

        if rec:
            records.append(rec)

    return records


def find_qdar_roots(base_dir: Path) -> list[Path]:
    """Find all potential QDAR directories containing workbooks or External/Internal folders."""
    if not base_dir.exists():
        return []

    # Check if base_dir has External/Internal subdirectories
    subdirs = [d for d in base_dir.iterdir() if d.is_dir()]
    if any(d.name.lower() in ("external", "internal") for d in subdirs):
        return [base_dir]

    qdar_candidates = []
    # Check for direct child named QDARS / QDR / QAD
    for child in subdirs:
        name_lower = child.name.lower()
        if "qdar" in name_lower or "qdr" in name_lower or "qad" in name_lower or "quality" in name_lower:
            qdar_candidates.append(child)

    if qdar_candidates:
        return qdar_candidates

    # Check deeper
    for child in subdirs:
        for grand in child.iterdir():
            if grand.is_dir() and grand.name.lower() in ("external", "internal"):
                return [child]

    return [base_dir]


def parse_all_qdar_files(qdar_dir: Path | None = None, year: int = 2025) -> list[dict]:
    all_records = []
    initial_dir = Path(qdar_dir) if qdar_dir else RAW_QDAR_DIR

    if not initial_dir.exists():
        log.error(f"QDAR directory not found: {initial_dir}")
        return all_records

    target_roots = find_qdar_roots(initial_dir)
    log.info(f"Using QDAR target roots: {[str(r) for r in target_roots]}")

    for target_dir in target_roots:
        ext_dir = target_dir / "External"
        if not ext_dir.exists():
            for d in target_dir.iterdir():
                if d.is_dir() and d.name.lower() == "external":
                    ext_dir = d
                    break

        if ext_dir.exists():
            ext_records = parse_qdar_directory(ext_dir, "QDR_EXTERNAL", year=year)
            all_records.extend(ext_records)
            log.info(f"  External: {len(ext_records)} records parsed from {ext_dir}")

        int_dir = target_dir / "Internal"
        if not int_dir.exists():
            for d in target_dir.iterdir():
                if d.is_dir() and d.name.lower() == "internal":
                    int_dir = d
                    break

        if int_dir.exists():
            int_records = parse_qdar_directory(int_dir, "QDR_INTERNAL", year=year)
            all_records.extend(int_records)
            log.info(f"  Internal: {len(int_records)} records parsed from {int_dir}")

        # Scan direct root files in target_dir that are not inside subdirectories
        direct_files = [f for f in target_dir.glob("*.xls*") if not f.name.startswith("~$") and f.is_file()]
        if direct_files and (ext_dir.exists() or int_dir.exists()):
            root_records = []
            cell_map = EXTERNAL_CELLS
            for fpath in direct_files:
                rec = parse_xlsx_qdar(fpath, cell_map, "QDR_EXTERNAL", year=year) or parse_xls_qdar(fpath, cell_map, "QDR_EXTERNAL", year=year)
                if rec:
                    root_records.append(rec)
            all_records.extend(root_records)
            log.info(f"  Direct Root Files: {len(root_records)} records parsed from {target_dir}")

        if not ext_dir.exists() and not int_dir.exists():
            records = parse_qdar_directory(target_dir, "QDR_EXTERNAL", year=year)
            all_records.extend(records)
            log.info(f"  Direct QDARs: {len(records)} records parsed from {target_dir}")

    return all_records


def main():
    log.info("=" * 60)
    log.info("QAD / QDAR FILE PARSER (WITH DEFECT LOGS) — Starting")
    log.info("=" * 60)

    all_records = parse_all_qdar_files()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(all_records, f, indent=2, ensure_ascii=False, default=str)

    log.info(f"Saved {len(all_records)} QDAR defect records to {OUTPUT_FILE}")
    log.info("QAD / QDAR FILE PARSER — Complete")


if __name__ == "__main__":
    main()
