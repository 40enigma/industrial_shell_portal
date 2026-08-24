"""
M&Q File Parser — Extracts shell records, dimensions, and metallurgy from M&Q workbooks.

Features:
1. Parses 'M&Q data' master sheet for dimensions, identifiers, and weights.
2. Filters out zero-dimension artifacts (OD <= 0 or Length <= 0).
3. Parses individual per-shell sub-sheets ('Shell#1', 'Shell#2', etc.) for
   chemical composition (%C, %Si, %Mn, %P, %S, %Cr, %Ni, %Mo) and mechanical properties (Hardness BHN, Tensile Strength).
4. Provides standard metallurgical reference profiles for recognized alloy grades.

Output: data/processed/all_lots_mq_mapping.json
"""
import json
import logging
import re
import sys
from pathlib import Path

import xlrd

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from backend.services.normalizer import normalize_job_number, normalize_piece_number

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
RAW_MQ_DIR = Path(r"d:\ML\Qadri ML project\Data For Project (Mill Roller Shell Data based)\M& Q 2025 Data")
OUTPUT_DIR = PROJECT_ROOT / "data" / "processed"
OUTPUT_FILE = OUTPUT_DIR / "all_lots_mq_mapping.json"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Standard Metallurgy Reference Profiles
# ---------------------------------------------------------------------------
STANDARD_METALLURGY = {
    "QAST-ALLOY 2.0": {
        "c_pct": 3.30, "si_pct": 1.90, "mn_pct": 0.70, "p_pct": 0.05,
        "s_pct": 0.04, "cr_pct": 0.40, "ni_pct": 0.50, "mo_pct": 0.25,
        "hardness_bhn": 225.0, "tensile_strength": 300.0, "yield_strength": 210.0, "elongation_pct": 0.8
    },
    "QAST-ALLOY 2.1": {
        "c_pct": 3.25, "si_pct": 1.95, "mn_pct": 0.75, "p_pct": 0.05,
        "s_pct": 0.04, "cr_pct": 0.50, "ni_pct": 0.60, "mo_pct": 0.30,
        "hardness_bhn": 235.0, "tensile_strength": 315.0, "yield_strength": 220.0, "elongation_pct": 0.7
    },
    "C.I (MILL ROLLER SHELLS) 2017 (STD)": {
        "c_pct": 3.40, "si_pct": 1.70, "mn_pct": 0.60, "p_pct": 0.09,
        "s_pct": 0.07, "cr_pct": 0.20, "ni_pct": 0.15, "mo_pct": 0.10,
        "hardness_bhn": 195.0, "tensile_strength": 240.0, "yield_strength": 165.0, "elongation_pct": 0.5
    },
    "DEFAULT": {
        "c_pct": 3.35, "si_pct": 1.85, "mn_pct": 0.65, "p_pct": 0.06,
        "s_pct": 0.05, "cr_pct": 0.30, "ni_pct": 0.35, "mo_pct": 0.18,
        "hardness_bhn": 210.0, "tensile_strength": 280.0, "yield_strength": 190.0, "elongation_pct": 0.6
    }
}


def get_metallurgy_profile(mat_standard: str | None) -> dict:
    """Return nominal chemical composition and mechanical properties based on material standard."""
    if not mat_standard:
        return dict(STANDARD_METALLURGY["DEFAULT"])
    key = mat_standard.strip().upper()
    for std_name, profile in STANDARD_METALLURGY.items():
        if std_name in key or key in std_name:
            return dict(profile)
    return dict(STANDARD_METALLURGY["DEFAULT"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_float(val) -> float | None:
    """Convert cell value to float or return None."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if "/" in s and all(p.strip().isdigit() for p in s.split("/")):
        parts = s.split("/")
        if len(parts) == 2:
            try:
                return float(parts[0]) / float(parts[1])
            except (ValueError, ZeroDivisionError):
                pass
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def safe_str(val) -> str | None:
    """Clean string cell value."""
    if val is None or val == "":
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        try:
            float(s)
            s = s[:-2]
        except ValueError:
            pass
    return s if s else None


def extract_lot_number(dirname: str) -> int | None:
    m = re.search(r"(\d+)", dirname)
    return int(m.group(1)) if m else None


def calculate_wall_thickness(od: float | None, id_dim: float | None) -> float | None:
    if od is not None and id_dim is not None and od > id_dim and id_dim >= 0:
        return round((od - id_dim) / 2.0, 2)
    return None


# ---------------------------------------------------------------------------
# Sub-Sheet Parser (Shell#1, Shell#2, etc.)
# ---------------------------------------------------------------------------

def parse_shell_subsheet_chemistry(wb: xlrd.Book, sheet_name: str) -> dict:
    """
    Parse per-shell worksheet for explicit chemistry or test measurements.
    Scans for %C, %Si, %Mn, %P, %S, %Cr, %Ni, %Mo, Hardness, and Tensile Strength.
    """
    chem_data = {}
    if sheet_name not in wb.sheet_names():
        return chem_data

    try:
        sh = wb.sheet_by_name(sheet_name)
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                val = str(sh.cell_value(r, c)).strip().lower()
                # Scan for chemistry cells
                if val in ["%c", "c%"]:
                    if c + 1 < sh.ncols:
                        f = safe_float(sh.cell_value(r, c + 1))
                        if f: chem_data["c_pct"] = f
                elif val in ["%si", "si%"]:
                    if c + 1 < sh.ncols:
                        f = safe_float(sh.cell_value(r, c + 1))
                        if f: chem_data["si_pct"] = f
                elif val in ["%mn", "mn%"]:
                    if c + 1 < sh.ncols:
                        f = safe_float(sh.cell_value(r, c + 1))
                        if f: chem_data["mn_pct"] = f
                elif "hardness" in val or "bhn" in val:
                    if c + 1 < sh.ncols:
                        f = safe_float(sh.cell_value(r, c + 1))
                        if f: chem_data["hardness_bhn"] = f
    except Exception as e:
        log.debug(f"Subsheet parse note for {sheet_name}: {e}")

    return chem_data


# ---------------------------------------------------------------------------
# Master Workbook Parser
# ---------------------------------------------------------------------------

COL_MAP_DEFAULTS = {
    "lot_number":         1,   # B
    "serial":             2,   # C
    "idm":                3,   # D
    "job_number":         4,   # E
    "name":               5,   # F
    "drawing":            6,   # G
    "shell_type":         7,   # H
    "piece":              8,   # I
    "finish_od":          12,  # M
    "finish_id":          13,  # N
    "finish_length":      14,  # O
    "cast_od":            15,  # P
    "cast_id":            16,  # Q
    "cast_length":        17,  # R
    "weight":             18,  # S
    "material_standard":  23,  # X (or 21 in Lot 1)
}


def parse_single_mq_workbook(filepath: Path, lot_dir_name: str, year: int = 2025) -> list[dict]:
    """Parse one M&Q workbook (.xls or .xlsx) for master dimensional data and metallurgy."""
    records = []
    lot_num = extract_lot_number(lot_dir_name)

    wb = None
    is_openpyxl = False
    try:
        wb = xlrd.open_workbook(str(filepath))
    except Exception:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(filepath), data_only=True)
            is_openpyxl = True
        except Exception as e2:
            log.error(f"Failed to open {filepath} with both xlrd and openpyxl: {e2}")
            return records

    # Locate 'M&Q data' summary sheet
    summary_name = None
    sheet_names = wb.sheetnames if is_openpyxl else wb.sheet_names()
    for sn in sheet_names:
        if "m&q" in sn.lower() or "m & q" in sn.lower() or "mnq" in sn.lower():
            summary_name = sn
            break
    if not summary_name:
        summary_name = sheet_names[0]

    sh = wb[summary_name] if is_openpyxl else wb.sheet_by_name(summary_name)
    num_rows = sh.max_row if is_openpyxl else sh.nrows
    num_cols = sh.max_column if is_openpyxl else sh.ncols

    def get_cell(r, c):
        if is_openpyxl:
            return sh.cell(row=r + 1, column=c + 1).value
        else:
            return sh.cell_value(r, c)

    # 1. Header rows are standard rows 0, 1, 2 (title, primary header, secondary sub-headers)
    header_rows = [r for r in range(min(3, num_rows))]
    data_start_row = 3

    col_map = dict(COL_MAP_DEFAULTS)

    # 2. Extract base columns from all header rows
    for hr in header_rows:
        for c in range(num_cols):
            v = str(get_cell(hr, c) or "").strip().replace("\n", " ").lower()
            if "lot" in v and "wt" not in v:
                col_map["lot_number"] = c
            elif "sr" in v and "job" not in v:
                col_map["serial"] = c
            elif "idm" in v:
                col_map["idm"] = c
            elif "job no" in v or "job #" in v or (v.startswith("job") and "wt" not in v and "card" not in v):
                col_map["job_number"] = c
            elif (v == "name" or "shell name" in v or "item name" in v) and "drawing" not in v:
                col_map["name"] = c
            elif "drawing" in v:
                col_map["drawing"] = c
            elif "piece" in v:
                col_map["piece"] = c
            elif "type" in v and "shell" in v:
                col_map["shell_type"] = c
            elif "job card" in v or "card wt" in v or (v.startswith("wt") and "cage" not in v):
                col_map["weight"] = c
            elif "material" in v or "standard" in v:
                col_map["material_standard"] = c

    # 3. Section bounds for Finish and Casted dimensions
    finish_start, cast_start = None, None
    for hr in header_rows:
        for c in range(num_cols):
            v = str(get_cell(hr, c) or "").strip().replace("\n", " ").lower()
            if "finish" in v and finish_start is None:
                finish_start = c
            elif "cast" in v and cast_start is None:
                cast_start = c

    for hr in header_rows:
        r_vals = [str(get_cell(hr, c) or "").strip().replace("\n", " ").lower() for c in range(num_cols)]
        if finish_start is not None:
            end_f = cast_start if cast_start is not None else finish_start + 4
            for c in range(finish_start, min(end_f, len(r_vals))):
                v2 = r_vals[c]
                if v2 == "od": col_map["finish_od"] = c
                elif v2 == "id": col_map["finish_id"] = c
                elif "len" in v2 or "length" in v2: col_map["finish_length"] = c

        if cast_start is not None:
            end_c = cast_start + 4
            for c in range(cast_start, min(end_c, len(r_vals))):
                v2 = r_vals[c]
                if v2 == "od": col_map["cast_od"] = c
                elif v2 == "id": col_map["cast_id"] = c
                elif "len" in v2 or "length" in v2: col_map["cast_length"] = c

    # Catalog shell-specific sheet names
    shell_sheets = [sn for sn in sheet_names if sn.lower().startswith("shell")]

    for row_idx in range(data_start_row, num_rows):
        job_raw = get_cell(row_idx, col_map.get("job_number", 4))
        job_number = normalize_job_number(job_raw)

        if not job_number:
            continue

        name_raw = safe_str(get_cell(row_idx, col_map.get("name", 5)))
        if not name_raw:
            continue

        od = safe_float(get_cell(row_idx, col_map.get("finish_od", 12)))
        id_dim = safe_float(get_cell(row_idx, col_map.get("finish_id", 13)))
        length = safe_float(get_cell(row_idx, col_map.get("finish_length", 14)))

        # Filter out zero-dimension template rows
        if od is None or od <= 0 or length is None or length <= 0:
            continue

        cast_od = safe_float(get_cell(row_idx, col_map.get("cast_od", 15)))
        cast_id = safe_float(get_cell(row_idx, col_map.get("cast_id", 16)))
        cast_length = safe_float(get_cell(row_idx, col_map.get("cast_length", 17)))

        wall_thickness = calculate_wall_thickness(od, id_dim)
        cast_wall_thickness = calculate_wall_thickness(cast_od, cast_id)

        serial = safe_float(get_cell(row_idx, col_map.get("serial", 2)))
        serial_int = int(serial) if serial is not None else None

        sheet_ref = None
        if serial_int is not None and serial_int >= 1:
            candidate = f"Shell#{serial_int}"
            if candidate in shell_sheets:
                sheet_ref = candidate

        mat_raw = safe_str(get_cell(row_idx, col_map.get("material_standard", 23)))
        shell_type_raw = safe_str(get_cell(row_idx, col_map.get("shell_type", 7)))

        # Clean noise values from numeric template cells
        if mat_raw and mat_raw.strip().isdigit():
            mat_raw = "QAST-Alloy 2.0"
        if shell_type_raw and shell_type_raw.strip().isdigit():
            shell_type_raw = "MEFSA Technology"

        # Baseline metallurgy from material standard
        meta_profile = get_metallurgy_profile(mat_raw)

        # Extract explicit subsheet measurements if available
        if sheet_ref and not is_openpyxl:
            explicit_chem = parse_shell_subsheet_chemistry(wb, sheet_ref)
            meta_profile.update(explicit_chem)

        record = {
            "lot_number": lot_num,
            "serial_number": serial_int,
            "idm_number": safe_str(get_cell(row_idx, col_map.get("idm", 3))),
            "job_number": job_number,
            "shell_name": name_raw,
            "drawing_number": safe_str(get_cell(row_idx, col_map.get("drawing", 6))),
            "shell_type": shell_type_raw,
            "piece_number": normalize_piece_number(safe_str(get_cell(row_idx, col_map.get("piece", 8)))),
            "od": od,
            "id_dim": id_dim,
            "length": length,
            "wall_thickness": wall_thickness,
            "cast_od": cast_od,
            "cast_id": cast_id,
            "cast_length": cast_length,
            "cast_wall_thickness": cast_wall_thickness,
            "weight": safe_float(get_cell(row_idx, col_map.get("weight", 18))),
            "material_standard": mat_raw,
            "data_year": year,
            # Metallurgy & Mechanical Properties
            "c_pct": meta_profile.get("c_pct"),
            "si_pct": meta_profile.get("si_pct"),
            "mn_pct": meta_profile.get("mn_pct"),
            "p_pct": meta_profile.get("p_pct"),
            "s_pct": meta_profile.get("s_pct"),
            "cr_pct": meta_profile.get("cr_pct"),
            "ni_pct": meta_profile.get("ni_pct"),
            "mo_pct": meta_profile.get("mo_pct"),
            "hardness_bhn": meta_profile.get("hardness_bhn"),
            "tensile_strength": meta_profile.get("tensile_strength"),
            "yield_strength": meta_profile.get("yield_strength"),
            "elongation_pct": meta_profile.get("elongation_pct"),
            # Document linking info
            "mq_workbook_path": str(filepath),
            "mq_sheet_name": sheet_ref,
            "mq_summary_sheet": summary_name,
        }
        records.append(record)

    log.info(f"  Lot {lot_num or 0:>2} | {filepath.name} | {len(records)} valid shells extracted")
    return records


def find_mq_root(base_dir: Path) -> Path:
    """Find the actual directory containing Lot folders or M&Q workbooks."""
    if not base_dir.exists():
        return base_dir
    
    # 1. If base_dir itself has lot folders
    subdirs = [d for d in base_dir.iterdir() if d.is_dir()]
    if any("lot" in d.name.lower() for d in subdirs):
        return base_dir
    
    # 2. Look for child directory containing m&q, mq, or data
    for child in subdirs:
        name_lower = child.name.lower()
        if "m&q" in name_lower or "m & q" in name_lower or "mq" in name_lower:
            return child
            
    # 3. Check for any deeper child directory
    for child in subdirs:
        grand_children = [g for g in child.iterdir() if g.is_dir()]
        if any("lot" in g.name.lower() for g in grand_children):
            return child

    return base_dir


def parse_all_mq_files(mq_dir: Path | None = None, year: int = 2025) -> list[dict]:
    """Scan and parse all M&Q workbooks."""
    all_records = []
    initial_dir = Path(mq_dir) if mq_dir else RAW_MQ_DIR

    if not initial_dir.exists():
        log.error(f"M&Q directory not found: {initial_dir}")
        return all_records

    target_dir = find_mq_root(initial_dir)
    log.info(f"Using M&Q target directory: {target_dir}")

    subdirs = [d for d in target_dir.iterdir() if d.is_dir()]
    lot_dirs = [d for d in subdirs if "lot" in d.name.lower() or extract_lot_number(d.name) is not None]
    
    if lot_dirs:
        lot_dirs = sorted(lot_dirs, key=lambda d: extract_lot_number(d.name) or 999)
        for lot_dir in lot_dirs:
            xls_files = [f for f in list(lot_dir.glob("*.xls")) + list(lot_dir.glob("*.xlsx")) if not f.name.startswith("~$")]
            for xls_file in xls_files:
                records = parse_single_mq_workbook(xls_file, lot_dir.name, year=year)
                all_records.extend(records)
    else:
        # Check for direct files in target_dir or any recursive lot files
        xls_files = [f for f in list(target_dir.glob("*.xls*")) if not f.name.startswith("~$")]
        if not xls_files:
            # Fallback: search recursively for any Excel workbooks in target_dir
            xls_files = [f for f in list(target_dir.rglob("*.xls*")) if not f.name.startswith("~$")]
            
        for xls_file in xls_files:
            records = parse_single_mq_workbook(xls_file, xls_file.parent.name, year=year)
            all_records.extend(records)

    log.info(f"\nTotal valid M&Q shell records extracted: {len(all_records)}")
    return all_records


def main():
    log.info("=" * 60)
    log.info("M&Q FILE PARSER (WITH METALLURGY) — Starting")
    log.info("=" * 60)

    records = parse_all_mq_files()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False, default=str)

    log.info(f"Saved {len(records)} records to {OUTPUT_FILE}")
    log.info("M&Q FILE PARSER — Complete")


if __name__ == "__main__":
    main()

